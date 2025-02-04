"""
OpenNEM AEMO facility closure dates parser.

"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Union

from openpyxl import load_workbook
from pydantic import ValidationError
from pydantic.class_validators import validator

from opennem.core.normalizers import is_number, normalize_duid
from opennem.schema.core import BaseConfig

logger = logging.getLogger("opennem.parsers.aemo_nem_facility_closures")

WORKBOOK_SHEET_NAME = "Expected Closure Year"

CLOSURE_SHEET_FIELDS = [
    "station_name",
    "duid",
    "expected_closure_year",
    "expected_closure_date",
]


def _clean_expected_closure_year(closure_year: Union[str, int]) -> Optional[int]:
    """Clean up expected closure year because sometimes they just put comments in the field"""
    if is_number(closure_year):
        return int(closure_year)

    return None


class AEMOClosureRecord(BaseConfig):
    station_name: str
    duid: Optional[str]
    expected_closure_year: Optional[int]
    expected_closure_date: Optional[datetime]

    _validate_closure_year = validator("expected_closure_year", pre=True)(
        _clean_expected_closure_year
    )

    _clean_duid = validator("duid", pre=True)(normalize_duid)


def parse_aemo_closures_xls() -> List[AEMOClosureRecord]:
    """Parse the AEMO NEM closures spreadsheet"""
    aemo_path = (
        Path(__file__).parent.parent.parent
        / "data"
        / "aemo"
        / "generating-unit-expected-closure-year.xlsx"
    )

    if not aemo_path.is_file():
        raise Exception("Not found: {}".format(aemo_path))

    # @TODO split here to read ByteIO from download / local file
    wb = load_workbook(aemo_path, data_only=True)

    generator_ws = wb[WORKBOOK_SHEET_NAME]

    records = []

    for row in generator_ws.iter_rows(min_row=2, values_only=True):
        row_collapsed = row[0:2] + row[3:5]

        return_dict = dict(zip(CLOSURE_SHEET_FIELDS, list(row_collapsed)))

        r = None

        try:
            r = AEMOClosureRecord(**return_dict)
        except ValidationError as e:
            logger.error("Validation error: {}. {}".format(e, return_dict))

        if r:
            records.append(r)

    return records


if __name__ == "__main__":
    p = parse_aemo_closures_xls()

    from pprint import pprint

    pprint(p)
