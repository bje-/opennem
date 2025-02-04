"""
OpenNEM NEM General Information Parser

This module parses the general information Excel files that are released
by AEMO (almost) every month. The downloads can be found on this page:


https://www.aemo.com.au/energy-systems/electricity/national-electricity-market-nem/nem-forecasting-and-planning/forecasting-and-planning-data/generation-information

In opennem.spiders.aemo.monitoring there is a spider that finds the new files
and they will be passed into this file to be output into a StationSet schema,
ready for manipulation, export into different formats - or, for our primary
purpose - to import new facilities and updates into the OpenNEM database.
"""

import logging
import re

# from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from pydantic import validator
from pydantic.types import conint
from sqlalchemy.sql import select

import opennem  # noqa: 401
from opennem.core.normalizers import (
    clean_capacity,
    clean_float,
    normalize_duid,
    station_name_cleaner,
)
from opennem.db import SessionLocal, get_database_engine
from opennem.db.models.opennem import Facility
from opennem.schema.core import BaseConfig

logger = logging.getLogger("opennem.parsers.aemo.gi")

GI_EXISTING_NEW_GEN_KEYS = {
    "region": "A",
    "AssetType": "B",
    "StationName": "C",
    "Owner": "D",
    "TechType": "E",
    "FuelType": "F",
    "duid": "G",
    "units_no": "H",
    "capacity_registered": "M",
    "StorageCapacity": "N",
    "UnitStatus": "O",
    "DispatchType": "P",
    "UseDate": "Q",
    "ClosureDateExpected": "R",
    "ClosureDate": "S",
    "SurveyID": "V",
    "FuelSummary": "U",
    "SurveyEffective": "Y",
}


def excel_column_to_column_index(excel_column: str) -> int:
    """Takes an excel column like 'A' and converts to column index"""
    return ord(excel_column.upper()) - 64


AEMO_GI_FUELTECH_MAP = {
    "Solar": "solar_utility",
    "Battery Storage": "battery_charging",
    "Coal": "coal_black",
    "CCGT": "gas_ccgt",
    "Water": "hydo",
    "Wind": "wind",
    "Biomass": "bioenergy_biomass",
    "OCGT": "gas_ocgt",
}

AEMO_GI_STATUS_MAP = {
    "Anticipated": None,
    "Committed": "committed",
    "Publicly Announced Upgrade": None,
    "Committed¹": "committed",
    "Committed Upgrade": None,
    "Withdrawn - Permanent": None,
    "Committed*": "committed",
    "In Service - Announced Withdrawal (Permanent)": "operating",
    "In Commissioning": "commissioning",
    "In Service": "operating",
    "Publicly Announced": "announced",
    "Anticipated Upgrade": None,
}


def aemo_gi_fueltech_to_fueltech(gi_fueltech: Optional[str]) -> Optional[str]:
    if not gi_fueltech:
        return None

    if gi_fueltech not in AEMO_GI_FUELTECH_MAP.keys():
        return None

    return AEMO_GI_FUELTECH_MAP[gi_fueltech]


def aemo_gi_status_map(gi_status: Optional[str]) -> Optional[str]:
    if not gi_status:
        return None

    if gi_status not in AEMO_GI_STATUS_MAP.keys():
        return None

    return AEMO_GI_STATUS_MAP[gi_status]


def aemo_gi_capacity_cleaner(cap: Optional[str]) -> Optional[float]:
    """Custom capacity cleaner because sometimes its parsed as silly
    text like a range (ie. '150 - 180'"""
    if isinstance(cap, int) or isinstance(cap, float):
        return cap

    if not cap:
        return None

    cap = cap.strip()

    num_part = re.search(r"^[\d\.]+", cap)

    if not num_part:
        return None

    num_extracted = str(num_part.group(0))

    num_extracted_and_clean = clean_float(num_extracted)

    return num_extracted_and_clean


class AEMOGIRecord(BaseConfig):
    name: str
    region: str
    fueltech_id: Optional[str]
    status_id: Optional[str]
    duid: Optional[str]
    units_no: Optional[int]
    capacity_registered: Optional[float]

    # expected_closure_year: Optional[int]
    # expected_closure_date: Optional[datetime]

    # _validate_closure_year = validator("expected_closure_year", pre=True)(
    #     _clean_expected_closure_year
    # )

    _clean_duid = validator("duid", pre=True, allow_reuse=True)(normalize_duid)
    _clean_capacity = validator("capacity_registered", pre=True, allow_reuse=True)(
        aemo_gi_capacity_cleaner
    )


def parse_aemo_general_information(filename: str) -> List[AEMOGIRecord]:
    wb = load_workbook(filename, data_only=True)

    SHEET_KEY = "ExistingGeneration&NewDevs"

    if SHEET_KEY not in wb:
        raise Exception("Doesn't look like a GI spreadsheet")

    ws = wb[SHEET_KEY]

    records = []

    for row in ws.iter_rows(min_row=3, values_only=True):

        # pick out the columns we want
        # lots of hidden columns in the sheet
        row_collapsed = [
            row[excel_column_to_column_index(i) - 1] for i in GI_EXISTING_NEW_GEN_KEYS.values()
        ]

        return_dict = dict(zip(GI_EXISTING_NEW_GEN_KEYS, list(row_collapsed)))

        # break at end of data records
        # GI has a blank line before garbage notes
        if row[0] is None:
            break

        if return_dict is None:
            raise Exception("Failed on row: {}".format(row))

        return_dict = {
            **return_dict,
            **{
                "name": station_name_cleaner(return_dict["StationName"]),
                "status_id": aemo_gi_status_map(return_dict["UnitStatus"]),
                "fueltech_id": aemo_gi_fueltech_to_fueltech(return_dict["FuelSummary"]),
            },
        }

        return_model = AEMOGIRecord(**return_dict)

        records.append(return_model)

    return records


def get_unique_values_for_field(records: List[Dict], field_name: str) -> List[Any]:
    return list(set([i[field_name] for i in records]))


def facility_matcher(records: List[AEMOGIRecord]) -> None:
    with SessionLocal() as sess:

        for gi_record in records:
            if not gi_record.duid:
                continue

            gi_lookup: Optional[Facility] = sess.execute(
                select(Facility).where(Facility.code == gi_record.duid)
            ).one_or_none()

            if not gi_lookup:
                logger.info(f"MISS: {gi_record.duid} {gi_record.name}")
                continue

            gi_db: Facility = gi_lookup[0]

            logger.info(
                f"HIT {gi_record.duid} {gi_record.name} - currently {gi_db.status_id} change to => {gi_record.status_id}"
            )

            gi_db.status_id = gi_record.status_id

            sess.add(gi_db)

        sess.commit()


# debug entrypoint
if __name__ == "__main__":
    aemo_gi_testfile = (
        Path(__file__).parent.parent.parent.parent / "data" / "aemo" / "nem_gi_202107.xlsx"
    )

    if not aemo_gi_testfile.is_file():
        print(f"file not found: {aemo_gi_testfile}")
        raise Exception("nahhh file")

    logger.info(f"Loading: {aemo_gi_testfile}")

    records = parse_aemo_general_information(str(aemo_gi_testfile))
    records = list(filter(lambda x: x.status_id in ["committed", "commissioning"], records))

    from pprint import pprint

    # pprint(records)
    facility_matcher(records=records)
    # pprint(get_unique_values_for_field(records, "Region"))
